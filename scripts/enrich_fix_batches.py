#!/usr/bin/env python3
"""Enrich sentence_fix_batch_*_input.json with current sentences and violation details."""

import json
from pathlib import Path
import openpyxl

REPO_ROOT = Path(__file__).parent.parent
VIOLATIONS_PATH = REPO_ROOT / "data" / "sentence_violations.json"
XLSX_PATH = REPO_ROOT / "data" / "words.xlsx"

def load_violations():
    """Return {(simplified, col): [(word, level), ...]}"""
    with open(VIOLATIONS_PATH, encoding="utf-8") as f:
        violations = json.load(f)
    result = {}
    for v in violations:
        key = (v["word"], v["col"])
        result[key] = v["violations"]
    return result

def load_sentences():
    """Return {simplified: {sentence_1: ..., translation_1: ..., ...}}"""
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    simplified_idx = headers.index("simplified")
    sent_keys = ["sentence_1", "translation_1", "sentence_2", "translation_2",
                 "sentence_3", "translation_3"]
    sent_indices = {k: headers.index(k) for k in sent_keys}

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        word = row[simplified_idx]
        if not word or word in result:
            continue
        result[word] = {k: row[idx] for k, idx in sent_indices.items()}
    return result

def main():
    violations = load_violations()
    sentences = load_sentences()

    for batch_num in range(1, 27):
        in_path = REPO_ROOT / "data" / f"sentence_fix_batch_{batch_num}_input.json"
        out_path = REPO_ROOT / "data" / f"sentence_fix_batch_{batch_num}_enriched.json"

        with open(in_path, encoding="utf-8") as f:
            words = json.load(f)

        enriched = []
        for w in words:
            simplified = w["simplified"]
            sents = sentences.get(simplified, {})
            entry = {**w}
            # Add current sentences
            for k in ["sentence_1", "translation_1", "sentence_2", "translation_2",
                      "sentence_3", "translation_3"]:
                entry[k] = sents.get(k)
            # Add violation details per column
            entry["violation_details"] = {}
            for col in w.get("fix_cols", []):
                viol = violations.get((simplified, col), [])
                if viol:
                    entry["violation_details"][col] = [f"{tok}(L{lvl})" for tok, lvl in viol]
            enriched.append(entry)

        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(enriched, f, ensure_ascii=False, indent=2)
        print(f"Wrote {out_path.name}")

if __name__ == "__main__":
    main()
