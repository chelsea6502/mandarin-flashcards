"""Apply L5 second-pass audit fixes (JSON format) to words.xlsx."""
import json
import glob
import os
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'words.xlsx')
FIX_DIR = os.path.join(ROOT, 'data', 'level_data')

wb = openpyxl.load_workbook(XLSX)
ws = wb.active

# Build column-name → column-index mapping from header row
header = [cell.value for cell in ws[1]]
col = {name: idx for idx, name in enumerate(header)}

# Load all pass2 fix files in order
fix_files = sorted(glob.glob(os.path.join(FIX_DIR, 'L5_chunk*_pass2_fixes.json')))
print(f"Found {len(fix_files)} fix files:")
for f in fix_files:
    print(f"  {os.path.basename(f)}")

all_fixes = []
for path in fix_files:
    with open(path, encoding='utf-8') as fh:
        fixes = json.load(fh)
    all_fixes.extend(fixes)

print(f"\nTotal fixes to apply: {len(all_fixes)}")

# Separate removals from edits
edits = [fx for fx in all_fixes if fx['field'] != 'REMOVE']
removals = [fx for fx in all_fixes if fx['field'] == 'REMOVE']

# Apply edits
applied = 0
skipped = 0
for fx in edits:
    row_num = fx['row']
    field = fx['field']
    old_val = fx['old']
    new_val = fx['new']

    if field not in col:
        print(f"  WARN: unknown field '{field}' in fix for row {row_num}")
        skipped += 1
        continue

    cell = ws.cell(row=row_num, column=col[field] + 1)  # openpyxl is 1-indexed
    if cell.value != old_val:
        print(f"  SKIP row {row_num} field {field}: expected {repr(old_val)}, got {repr(cell.value)}")
        skipped += 1
        continue

    cell.value = new_val
    applied += 1

# Apply removals (delete rows in reverse order)
removed = 0
rows_to_delete = sorted({fx['row'] for fx in removals}, reverse=True)
for row_num in rows_to_delete:
    ws.delete_rows(row_num)
    removed += 1

wb.save(XLSX)
print(f"\nEdits applied : {applied}")
print(f"Edits skipped : {skipped}")
print(f"Rows removed  : {removed}")
print(f"Total changes : {applied + removed}")
