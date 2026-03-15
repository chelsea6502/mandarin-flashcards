#!/usr/bin/env python3
"""Merge missing_batch_*.json files into data/words.xlsx.

Each batch file is a JSON array of row dicts with keys:
  simplified, pinyin, traditional, pos, classifier, definition, new_hsk,
  sentence_1, translation_1, sentence_2, translation_2, sentence_3, translation_3

Words are inserted after the last existing row that shares the same new_hsk level,
so the sheet stays grouped by level.

Usage: python3 scripts/merge_missing_words.py
"""

import json
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).parent.parent
XLSX_PATH = REPO_ROOT / "data" / "words.xlsx"
BATCH_GLOB = "missing_batch_*.json"

FIELDS = [
    "simplified", "pinyin", "traditional", "pos", "classifier",
    "definition", "new_hsk",
    "sentence_1", "translation_1",
    "sentence_2", "translation_2",
    "sentence_3", "translation_3",
]

MERGE_COLS = ["simplified", "pinyin", "traditional", "pos", "classifier", "new_hsk"]


def load_batches() -> list[dict]:
    """Load and merge all batch JSON files, sorted by batch number."""
    batch_files = sorted(
        REPO_ROOT.glob(f"data/{BATCH_GLOB}"),
        key=lambda p: int(re.search(r"\d+", p.stem).group()),
    )
    rows = []
    for path in batch_files:
        with open(path, encoding="utf-8") as f:
            batch = json.load(f)
        print(f"  Loaded {len(batch)} rows from {path.name}")
        rows.extend(batch)
    return rows


def level_sort_key(level_str) -> int:
    """Convert 'L1'..'L6' to int for sorting; unknown → 99."""
    if isinstance(level_str, str) and level_str.startswith("L"):
        try:
            return int(level_str[1:])
        except ValueError:
            pass
    return 99


def unmerge_all(ws):
    """Remove all merge regions so we can safely insert rows."""
    for merge in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge))


def remerge(ws, headers: list[str]):
    """Re-merge non-definition columns for consecutive rows of the same word."""
    merge_col_indices = [headers.index(c) + 1 for c in MERGE_COLS if c in headers]
    simplified_idx = headers.index("simplified") + 1

    n_rows = ws.max_row
    i = 2  # data starts at row 2
    while i <= n_rows:
        word = ws.cell(i, simplified_idx).value
        if not word:
            i += 1
            continue
        j = i + 1
        while j <= n_rows and ws.cell(j, simplified_idx).value == word:
            j += 1
        if j > i + 1:
            for col in merge_col_indices:
                ws.merge_cells(
                    start_row=i, start_column=col,
                    end_row=j - 1, end_column=col,
                )
        i = j


def main():
    print("Loading batch files...")
    new_rows = load_batches()
    print(f"Total new rows: {len(new_rows)}")

    # Validate all required fields are present
    missing_fields = set()
    for row in new_rows:
        for f in FIELDS:
            if f not in row:
                missing_fields.add(f)
    if missing_fields:
        print(f"WARNING: some rows missing fields: {missing_fields}")

    # Normalise: ensure every field exists (fill with empty string if absent)
    for row in new_rows:
        for f in FIELDS:
            if f not in row:
                row[f] = ""

    print(f"Loading {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    # Check headers match
    for f in FIELDS:
        if f not in headers:
            raise ValueError(f"Field '{f}' not found in xlsx headers: {headers}")

    # Collect existing simplified words to avoid duplicates
    simplified_idx = headers.index("simplified")
    existing_words = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[simplified_idx]:
            existing_words.add(row[simplified_idx])

    # Filter out words already in xlsx
    to_add = [r for r in new_rows if r["simplified"] not in existing_words]
    skipped = len(new_rows) - len(to_add)
    if skipped:
        print(f"Skipping {skipped} rows already in xlsx.")
    print(f"Adding {len(to_add)} new rows.")

    if not to_add:
        print("Nothing to add.")
        return

    # Read all existing data rows (below header)
    existing_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        existing_data.append(list(row))

    # Unmerge before manipulation
    unmerge_all(ws)

    # Combine existing + new, sort by level then simplified
    new_hsk_idx = headers.index("new_hsk")
    for row in to_add:
        existing_data.append([row.get(f, "") for f in headers])

    existing_data.sort(key=lambda r: (level_sort_key(r[new_hsk_idx]), r[simplified_idx] or ""))

    # Rewrite all data rows
    # Delete all existing data rows first
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row_data in existing_data:
        ws.append(row_data)

    # Re-merge
    remerge(ws, headers)

    wb.save(XLSX_PATH)
    print(f"Saved {XLSX_PATH} ({ws.max_row - 1} data rows total).")


if __name__ == "__main__":
    main()
