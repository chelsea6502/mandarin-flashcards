#!/usr/bin/env python3
"""Compare words.xlsx against official HSK vocabulary source.

Usage: python3 scripts/check_words_coverage.py
"""

import json
from pathlib import Path
import openpyxl

REPO_ROOT = Path(__file__).parent.parent
XLSX_PATH = REPO_ROOT / "data" / "words.xlsx"
SOURCE_PATH = Path("/Users/chelsea/hsk-2025-data/vocabulary.json")

LEVEL_MAP = {
    "一级": 1, "二级": 2, "三级": 3, "四级": 4, "五级": 5, "六级": 6,
}
MAX_LEVEL = 6


import re as _re

def level_name_to_int(name: str) -> int | None:
    """Extract the primary (first) level from names like '一级（二级）（四级）'."""
    m = _re.search(r'[一二三四五六]级', name)
    if not m:
        return None
    return LEVEL_MAP.get(m.group())


def load_source() -> dict:
    """Load official HSK source, returning {word: level_int}."""
    with open(SOURCE_PATH, encoding="utf-8") as f:
        records = json.load(f)
    result = {}
    for r in records:
        level = level_name_to_int(r.get("levelName", ""))
        if level is None:
            continue  # skip levels 7-9 and unknown
        # Strip trailing digit from homographs (e.g. "两1" -> "两")
        word = r["word"].rstrip("0123456789")
        if word not in result:
            result[word] = level
        else:
            # Use the lower level for duplicates
            result[word] = min(result[word], level)
    return result


def load_xlsx() -> dict:
    """Load words.xlsx, returning {simplified: new_hsk_level_int}."""
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    simplified_idx = headers.index("simplified")
    new_hsk_idx = headers.index("new_hsk")
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        word = row[simplified_idx]
        level = row[new_hsk_idx]
        if word and word not in result:
            if isinstance(level, str) and level.startswith("L"):
                try:
                    result[word] = int(level[1:])
                except ValueError:
                    result[word] = None
            elif level is not None:
                try:
                    result[word] = int(level)
                except (ValueError, TypeError):
                    result[word] = None
            else:
                result[word] = None
    return result


def main():
    source = load_source()
    xlsx = load_xlsx()

    source_words = set(source.keys())
    xlsx_words = set(xlsx.keys())

    missing = source_words - xlsx_words
    extra = xlsx_words - source_words
    common = source_words & xlsx_words

    wrong_level = []
    for word in sorted(common):
        src_lvl = source[word]
        xlsx_lvl = xlsx[word]
        if src_lvl != xlsx_lvl:
            wrong_level.append((word, xlsx_lvl, src_lvl))

    print(f"Source words:     {len(source_words):5d}")
    print(f"XLSX words:       {len(xlsx_words):5d}")
    print(f"Missing from xlsx:{len(missing):5d}")
    print(f"Extra in xlsx:    {len(extra):5d}")
    print(f"Wrong level:      {len(wrong_level):5d}")
    print()

    if missing:
        print("=== MISSING FROM XLSX (first 50) ===")
        for w in sorted(missing)[:50]:
            print(f"  {w}  (source L{source[w]})")
        if len(missing) > 50:
            print(f"  ... and {len(missing) - 50} more")
        print()

    if wrong_level:
        print("=== WRONG LEVEL (first 50) ===")
        print(f"  {'word':<12} {'xlsx':>6} {'source':>8}")
        for word, got, want in sorted(wrong_level)[:50]:
            print(f"  {word:<12} L{got!s:>5} L{want!s:>7}")
        if len(wrong_level) > 50:
            print(f"  ... and {len(wrong_level) - 50} more")
        print()

    if extra:
        print("=== EXTRA IN XLSX (not in official source, first 20) ===")
        for w in sorted(extra)[:20]:
            print(f"  {w}  (xlsx L{xlsx[w]})")
        if len(extra) > 20:
            print(f"  ... and {len(extra) - 20} more")


if __name__ == "__main__":
    main()
