#!/usr/bin/env python3
"""Apply sentence fix batches to data/words.xlsx.

Reads all sentence_fix_batch_*_output.json files and updates the relevant
sentence/translation columns for each word in words.xlsx.

Usage: python3 scripts/apply_sentence_fixes.py
"""

import json
import re
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).parent.parent
XLSX_PATH = REPO_ROOT / "data" / "words.xlsx"


def load_all_fixes() -> dict:
    """Return {simplified: {col: value}} from all output batch files."""
    fixes = {}
    batch_files = sorted(
        REPO_ROOT.glob("data/sentence_fix_batch_*_output.json"),
        key=lambda p: int(re.search(r"\d+", p.stem).group()),
    )
    if not batch_files:
        print("No output batch files found.")
        return fixes
    for path in batch_files:
        with open(path, encoding="utf-8") as f:
            entries = json.load(f)
        for entry in entries:
            word = entry["simplified"]
            if word not in fixes:
                fixes[word] = {}
            for k, v in entry.items():
                if k != "simplified":
                    fixes[word][k] = v
        print(f"  Loaded {len(entries)} entries from {path.name}")
    return fixes


def main():
    print("Loading fix data...")
    fixes = load_all_fixes()
    print(f"Total words to fix: {len(fixes)}")

    if not fixes:
        return

    print(f"Loading {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    simplified_idx = headers.index("simplified") + 1
    col_indices = {h: i + 1 for i, h in enumerate(headers)}

    updated = 0
    for row in ws.iter_rows(min_row=2):
        word = row[simplified_idx - 1].value
        if not word or word not in fixes:
            continue
        for col_name, new_value in fixes[word].items():
            if col_name in col_indices:
                row[col_indices[col_name] - 1].value = new_value
                updated += 1

    wb.save(XLSX_PATH)
    print(f"Updated {updated} cells across {len(fixes)} words.")
    print(f"Saved {XLSX_PATH}")


if __name__ == "__main__":
    main()
