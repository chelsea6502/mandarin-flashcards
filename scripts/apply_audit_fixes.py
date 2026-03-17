#!/usr/bin/env python3
"""Apply audit fixes from a JSON file to words.xlsx."""

import json
import sys
from pathlib import Path
import openpyxl

WORDS_PATH = Path(__file__).parent.parent / "words.xlsx"

FIELD_TO_COL = {
    "pinyin": "pinyin",
    "traditional": "traditional",
    "pos": "pos",
    "classifier": "classifier",
    "definition": "definition",
    "sentence_1": "sentence_1",
    "translation_1": "translation_1",
    "sentence_2": "sentence_2",
    "translation_2": "translation_2",
    "sentence_3": "sentence_3",
    "translation_3": "translation_3",
}


def load_fixes(path: str) -> list:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def apply_fixes(fixes: list) -> tuple[int, int]:
    wb = openpyxl.load_workbook(WORDS_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col_index = {h: i + 1 for i, h in enumerate(headers) if h}

    simplified_col = col_index["simplified"]
    definition_col = col_index["definition"]

    corrections = 0
    removals = 0
    not_found = []

    rows_to_delete = []

    for fix in fixes:
        simplified = fix["simplified"]
        definition = fix.get("definition")
        field = fix["field"]
        old_val = fix.get("old")
        new_val = fix.get("new")

        matched_row = None
        # Collect all rows matching simplified
        candidates = []
        for row in ws.iter_rows(min_row=2):
            s_val = row[simplified_col - 1].value
            if s_val == simplified:
                candidates.append(row)

        if definition:
            # If definition provided, match on both
            for row in candidates:
                d_val = row[definition_col - 1].value
                if d_val == definition:
                    matched_row = row
                    break
            # Fallback: if definition match failed (e.g. a prior fix changed
            # the definition), try matching by old_val on the target field
            if matched_row is None and field in col_index and old_val is not None:
                target_col = col_index[field]
                for row in candidates:
                    if row[target_col - 1].value == old_val:
                        matched_row = row
                        break
        elif len(candidates) == 1:
            matched_row = candidates[0]
        elif field != "REMOVE_ROW" and field in col_index and old_val is not None:
            # Disambiguate by checking current field value matches old
            target_col = col_index[field]
            for row in candidates:
                if row[target_col - 1].value == old_val:
                    matched_row = row
                    break
        elif field == "REMOVE_ROW":
            # For removals without definition, take first candidate
            if candidates:
                matched_row = candidates[0]

        label = f"{simplified!r} / {definition!r}" if definition else f"{simplified!r}"
        if matched_row is None:
            not_found.append(f"  NOT FOUND: {label}")
            continue

        if field == "REMOVE_ROW":
            rows_to_delete.append(matched_row[0].row)
            removals += 1
            print(f"  REMOVE: {label}")
            continue

        if field not in col_index:
            print(f"  SKIP unknown field {field!r} for {simplified!r}")
            continue

        target_col = col_index[field]
        cell = matched_row[target_col - 1]
        current = cell.value

        # Treat None and null equivalently
        if current == old_val or (current is None and old_val is None):
            cell.value = new_val
            corrections += 1
            print(f"  FIX: {simplified!r} [{field}]: {old_val!r} -> {new_val!r}")
        else:
            print(f"  MISMATCH: {simplified!r} [{field}] expected {old_val!r}, got {current!r}")

    # Delete rows in reverse order to preserve row numbers
    for row_num in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_num)

    wb.save(WORDS_PATH)

    if not_found:
        print("\nNot found:")
        for msg in not_found:
            print(msg)

    return corrections, removals


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/apply_audit_fixes.py <fixes.json>")
        sys.exit(1)

    fixes_path = sys.argv[1]
    fixes = load_fixes(fixes_path)
    print(f"Loaded {len(fixes)} fixes from {fixes_path}")

    corrections, removals = apply_fixes(fixes)
    print(f"\nDone: {corrections} corrections, {removals} row removals")


if __name__ == "__main__":
    main()
