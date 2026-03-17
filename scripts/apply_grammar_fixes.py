#!/usr/bin/env python3
"""Apply audit fixes from a JSON file to grammar.xlsx."""

import json
import sys
from pathlib import Path
import openpyxl

GRAMMAR_PATH = Path(__file__).parent.parent / "grammar.xlsx"

FIELD_TO_COL = {
    "sentence_1": "sentence_1",
    "translation_1": "translation_1",
    "sentence_2": "sentence_2",
    "translation_2": "translation_2",
    "sentence_3": "sentence_3",
    "translation_3": "translation_3",
}


def load_fixes(path: str) -> list:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def apply_fixes(fixes: list) -> tuple[int, int]:
    wb = openpyxl.load_workbook(GRAMMAR_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col_index = {h: i + 1 for i, h in enumerate(headers) if h}

    content_col = col_index["content"]
    level_col = col_index["hsk_level"]

    corrections = 0
    not_found = []

    for fix in fixes:
        content = fix["content"]
        hsk_level = fix.get("hsk_level")
        field = fix["field"]
        old_val = fix.get("old")
        new_val = fix.get("new")

        if field not in FIELD_TO_COL:
            print(f"  SKIP unknown field {field!r} for {content!r}")
            continue

        matched_row = None
        candidates = []
        for row in ws.iter_rows(min_row=2):
            c_val = row[content_col - 1].value
            if c_val == content:
                candidates.append(row)

        if hsk_level:
            for row in candidates:
                l_val = row[level_col - 1].value
                if l_val == hsk_level:
                    matched_row = row
                    break

        # Fallback: match by old value on target field
        if matched_row is None and old_val is not None:
            target_col = col_index[field]
            for row in candidates:
                if row[target_col - 1].value == old_val:
                    matched_row = row
                    break

        if matched_row is None and len(candidates) == 1:
            matched_row = candidates[0]

        label = f"{content!r} ({hsk_level})" if hsk_level else f"{content!r}"
        if matched_row is None:
            not_found.append(f"  NOT FOUND: {label}")
            continue

        target_col = col_index[field]
        cell = matched_row[target_col - 1]
        current = cell.value

        if current == old_val or (current is None and old_val is None):
            cell.value = new_val
            corrections += 1
            print(f"  FIX: {label} [{field}]: {old_val!r} -> {new_val!r}")
        else:
            print(f"  MISMATCH: {label} [{field}] expected {old_val!r}, got {current!r}")

    wb.save(GRAMMAR_PATH)

    if not_found:
        print("\nNot found:")
        for msg in not_found:
            print(msg)

    return corrections, 0


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/apply_grammar_fixes.py <fixes.json>")
        sys.exit(1)

    fixes_path = sys.argv[1]
    fixes = load_fixes(fixes_path)
    print(f"Loaded {len(fixes)} fixes from {fixes_path}")

    corrections, _ = apply_fixes(fixes)
    print(f"\nDone: {corrections} corrections")


if __name__ == "__main__":
    main()
