#!/usr/bin/env python3
"""Verify that example sentences only use vocabulary at or below the row's HSK level.

Uses forward maximum matching (FMM) to segment sentences against the HSK word list.
Reports all violations grouped by severity.

Usage: python3 scripts/check_sentence_levels.py
"""

import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).parent.parent
XLSX_PATH = REPO_ROOT / "data" / "words.xlsx"
SOURCE_PATH = Path("/Users/chelsea/hsk-2025-data/vocabulary.json")

LEVEL_MAP = {"一级": 1, "二级": 2, "三级": 3, "四级": 4, "五级": 5, "六级": 6}


def level_name_to_int(name: str):
    m = re.search(r"[一二三四五六]级", name)
    return LEVEL_MAP.get(m.group()) if m else None


def load_hsk_vocab() -> dict:
    """Return {word: min_level} for all L1–6 words."""
    with open(SOURCE_PATH, encoding="utf-8") as f:
        records = json.load(f)
    vocab = {}
    for r in records:
        level = level_name_to_int(r.get("levelName", ""))
        if level is None:
            continue
        word = r["word"].rstrip("0123456789")
        if word not in vocab:
            vocab[word] = level
        else:
            vocab[word] = min(vocab[word], level)
    return vocab


def find_violations(text: str, row_level: int, vocab: dict, vocab_by_len: list) -> list[tuple[str, int]]:
    """Return list of (word, level) for all HSK words above row_level found in text.

    Uses simple substring matching. Suppresses a match if all its characters
    are already covered by a longer word at or below row_level.
    """
    if not text:
        return []

    # Find all HSK words present in the sentence, longest first
    found: list[tuple[str, int]] = []  # (word, level)
    for word in vocab_by_len:
        if word in text:
            found.append((word, vocab[word]))

    # Build coverage map: char_index -> min level of any word covering it
    # (only words at or below row_level count as "coverage")
    coverage = {}
    for word, level in found:
        if level <= row_level:
            start = 0
            while True:
                idx = text.find(word, start)
                if idx == -1:
                    break
                for i in range(idx, idx + len(word)):
                    coverage[i] = min(coverage.get(i, 99), level)
                start = idx + 1

    # Violations: words above row_level whose characters aren't all covered
    violations = []
    seen = set()
    for word, level in found:
        if level <= row_level or word in seen:
            continue
        seen.add(word)
        start = 0
        while True:
            idx = text.find(word, start)
            if idx == -1:
                break
            # Check if any character in this match is not covered
            uncovered = any(coverage.get(idx + i, 99) > row_level for i in range(len(word)))
            if uncovered:
                violations.append((word, level))
                break
            start = idx + 1

    return violations


def parse_xlsx_level(val) -> int | None:
    if isinstance(val, str) and val.startswith("L"):
        try:
            return int(val[1:])
        except ValueError:
            pass
    if val is not None:
        try:
            return int(val)
        except (ValueError, TypeError):
            pass
    return None


def main():
    print("Loading HSK vocabulary...")
    vocab = load_hsk_vocab()
    print(f"  {len(vocab)} unique words in L1–6")

    # Pre-sort vocab words longest-first for coverage checks
    vocab_by_len = sorted(vocab.keys(), key=len, reverse=True)

    print(f"Loading {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    simplified_idx = headers.index("simplified")
    new_hsk_idx = headers.index("new_hsk")
    sent_cols = {
        "sentence_1": headers.index("sentence_1"),
        "sentence_2": headers.index("sentence_2"),
        "sentence_3": headers.index("sentence_3"),
    }

    violations = []
    checked = 0
    skipped_no_level = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        simplified = row[simplified_idx]
        if not simplified:
            continue
        row_level = parse_xlsx_level(row[new_hsk_idx])
        if row_level is None:
            skipped_no_level += 1
            continue

        for col_name, col_idx in sent_cols.items():
            sentence = row[col_idx]
            if not sentence:
                continue
            checked += 1
            bad = find_violations(str(sentence), row_level, vocab, vocab_by_len)
            if bad:
                violations.append({
                    "word": simplified,
                    "row_level": row_level,
                    "col": col_name,
                    "sentence": sentence,
                    "violations": bad,
                })

    print(f"\nChecked {checked} sentences ({skipped_no_level} rows skipped — no level).")
    print(f"Violations found: {len(violations)}\n")

    if not violations:
        print("All sentences pass.")
        return

    # Group by severity (max level gap)
    def severity(v):
        return max(lvl - v["row_level"] for _, lvl in v["violations"])

    violations.sort(key=lambda v: (-severity(v), v["row_level"], v["word"]))

    # Save full violations list for repair use
    violations_path = REPO_ROOT / "data" / "sentence_violations.json"
    with open(violations_path, "w", encoding="utf-8") as f:
        json.dump(violations, f, ensure_ascii=False, indent=2)
    print(f"Full list saved to {violations_path}\n")

    # Summary by level
    print("=== SUMMARY BY ROW LEVEL ===")
    by_level = defaultdict(int)
    for v in violations:
        by_level[v["row_level"]] += 1
    for lvl in sorted(by_level):
        print(f"  L{lvl}: {by_level[lvl]} violations")

    # Most common violating words
    print("\n=== MOST COMMON VIOLATING WORDS ===")
    word_counts = defaultdict(int)
    for v in violations:
        for tok, lvl in v["violations"]:
            word_counts[(tok, lvl)] += 1
    top = sorted(word_counts.items(), key=lambda x: -x[1])[:20]
    print(f"  {'word':12} {'level':6} {'occurrences':>12}")
    for (tok, lvl), cnt in top:
        print(f"  {tok:12} L{lvl:<5} {cnt:>12}")

    # Print all violations
    print("\n=== ALL VIOLATIONS ===")
    prev_word = None
    for v in violations:
        if v["word"] != prev_word:
            print(f"\n  [{v['word']}]  row=L{v['row_level']}")
            prev_word = v["word"]
        bad_str = ", ".join(f"{tok}(L{lvl})" for tok, lvl in v["violations"])
        print(f"    {v['col']}: {v['sentence']}")
        print(f"      ^ {bad_str}")


if __name__ == "__main__":
    main()
