"""
validate_sentences.py

Check that each example sentence in words.xlsx only uses vocabulary
at or below the HSK level of the word being illustrated.

For each sentence, scans for any above-level HSK word appearing as a substring.
No tokenization needed — the HSK vocab list is the dictionary.

Usage:
    python3 validate_sentences.py [--xlsx data/words.xlsx] [--out violations.csv]

Output CSV columns:
    row, simplified, target_level, sentence_col, sentence,
    offending_word, word_level
"""

import argparse
import csv
from pathlib import Path

import openpyxl

LEVEL_ORDER = ["L1", "L2", "L3", "L4", "L5", "L6"]


def load_hsk_vocab(data_dir: Path) -> dict[str, str]:
    """Return {word: level} e.g. {'爱': 'L1', '结果': 'L3'}."""
    vocab: dict[str, str] = {}
    for level in LEVEL_ORDER:
        level_file = data_dir / f"hsk_level{level[1:]}.txt"
        if not level_file.exists():
            continue
        for line in level_file.read_text(encoding="utf-8").splitlines():
            word = line.strip()
            if word and word not in vocab:
                vocab[word] = level
    return vocab


def level_int(level: str) -> int:
    try:
        return LEVEL_ORDER.index(level)
    except ValueError:
        return 999


def check_sentence(sentence: str, target_level: str, above_level_words: list[tuple[str, str]]) -> list[tuple[str, str]]:
    """Return list of (word, level) for any above-level HSK words found in the sentence.
    Single-character words are skipped — they appear as components of longer words too
    often to be useful as substring matches.
    """
    return [(word, lvl) for word, lvl in above_level_words if len(word) >= 2 and word in sentence]


def run(xlsx_path: Path, out_path: Path, data_dir: Path) -> None:
    print(f"Loading HSK vocab from {data_dir}...", flush=True)
    vocab = load_hsk_vocab(data_dir)
    print(f"  {len(vocab)} words loaded", flush=True)

    # Pre-group words by level so we can build above-level sets per target level
    words_by_level: dict[str, list[tuple[str, str]]] = {lvl: [] for lvl in LEVEL_ORDER}
    for word, lvl in vocab.items():
        words_by_level[lvl].append((word, lvl))

    # For each target level, the words that are above it
    above: dict[str, list[tuple[str, str]]] = {}
    for i, lvl in enumerate(LEVEL_ORDER):
        above[lvl] = [
            (word, wlvl)
            for wlvl in LEVEL_ORDER[i + 1:]
            for word, _ in words_by_level[wlvl]
        ]

    print(f"Reading {xlsx_path}...", flush=True)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    sentence_cols = ["sentence_1", "sentence_2", "sentence_3"]
    violation_rows: list[dict] = []
    rows_checked = 0
    current_simplified = None
    current_level = None

    for row_num in range(2, ws.max_row + 1):
        v = ws.cell(row_num, col["simplified"]).value
        if v is not None:
            current_simplified = v
        v = ws.cell(row_num, col["new_hsk"]).value
        if v is not None:
            current_level = v

        if current_level not in LEVEL_ORDER:
            continue

        rows_checked += 1

        for scol in sentence_cols:
            sentence = ws.cell(row_num, col[scol]).value
            if not sentence:
                continue
            for word, word_level in check_sentence(sentence, current_level, above[current_level]):
                violation_rows.append({
                    "row": row_num,
                    "simplified": current_simplified,
                    "target_level": current_level,
                    "sentence_col": scol,
                    "sentence": sentence,
                    "offending_word": word,
                    "word_level": word_level,
                })

        if row_num % 500 == 0:
            print(f"  ...row {row_num}/{ws.max_row}", flush=True)

    print(f"\nChecked {rows_checked} rows.", flush=True)
    print(f"Found {len(violation_rows)} violations.", flush=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["row", "simplified", "target_level", "sentence_col",
                        "sentence", "offending_word", "word_level"],
        )
        writer.writeheader()
        writer.writerows(violation_rows)

    print(f"Written to {out_path}", flush=True)

    by_level: dict[str, int] = {}
    for v in violation_rows:
        by_level[v["target_level"]] = by_level.get(v["target_level"], 0) + 1
    if by_level:
        print("\nViolations by target level:")
        for lvl in LEVEL_ORDER:
            if lvl in by_level:
                print(f"  {lvl}: {by_level[lvl]}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate HSK sentence vocabulary levels")
    parser.add_argument("--xlsx", default="data/words.xlsx")
    parser.add_argument("--out", default="data/violations.csv")
    parser.add_argument("--data-dir", default="data")
    args = parser.parse_args()
    run(Path(args.xlsx), Path(args.out), Path(args.data_dir))


if __name__ == "__main__":
    main()
