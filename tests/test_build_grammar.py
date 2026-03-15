import sys
import pytest
import openpyxl
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from scripts.build_grammar import parse_grammar_points, write_xlsx


SAMPLE_L1 = """\
A.1 一级语法点
A.1.1 词类
A.1.1.1 名词
【一01】方位名词：上、下、里

书在桌子上。
手机在书包里。
房间里没有人。

A.1.1.2 动词
【一02】能愿动词：会、能
我不会说中文。
明天你能来吗？
"""


def test_parse_extracts_ids():
    rows = parse_grammar_points(SAMPLE_L1, "L1")
    assert len(rows) == 2
    assert rows[0]["id"] == "一01"
    assert rows[1]["id"] == "一02"


def test_parse_extracts_names():
    rows = parse_grammar_points(SAMPLE_L1, "L1")
    assert rows[0]["name"] == "方位名词：上、下、里"
    assert rows[1]["name"] == "能愿动词：会、能"


def test_parse_extracts_sentences():
    rows = parse_grammar_points(SAMPLE_L1, "L1")
    assert rows[0]["sentence_1"] == "书在桌子上。"
    assert rows[0]["sentence_2"] == "手机在书包里。"
    assert rows[0]["sentence_3"] == "房间里没有人。"


def test_parse_skips_phrase_lists():
    """Bare phrase lists without sentence-final punctuation are skipped."""
    text = "【一01】方位名词：上\n\n桌子上 树下\n书在桌子上。\n"
    rows = parse_grammar_points(text, "L1")
    assert rows[0]["sentence_1"] == "书在桌子上。"


def test_parse_fewer_than_three_sentences():
    """Only 2 sentences available — sentence_3 is empty."""
    rows = parse_grammar_points(SAMPLE_L1, "L1")
    # 【一02】 block has exactly 2 sentences
    assert rows[1]["sentence_1"] == "我不会说中文。"
    assert rows[1]["sentence_2"] == "明天你能来吗？"
    assert rows[1]["sentence_3"] == ""


def test_parse_sets_hsk_level():
    rows = parse_grammar_points(SAMPLE_L1, "L3")
    assert rows[0]["hsk_level"] == "L3"
    assert rows[1]["hsk_level"] == "L3"


def test_parse_description_is_empty():
    rows = parse_grammar_points(SAMPLE_L1, "L1")
    assert rows[0]["description"] == ""


def test_parse_translations_are_empty():
    rows = parse_grammar_points(SAMPLE_L1, "L1")
    for row in rows:
        assert row["translation_1"] == ""
        assert row["translation_2"] == ""
        assert row["translation_3"] == ""


GRAMMAR_ROW = {
    "id": "一01",
    "name": "方位名词：上、下、里",
    "hsk_level": "L1",
    "description": "",
    "sentence_1": "书在桌子上。",
    "translation_1": "",
    "sentence_2": "手机在书包里。",
    "translation_2": "",
    "sentence_3": "",
    "translation_3": "",
}


def test_write_xlsx_creates_file(tmp_path):
    out = tmp_path / "grammar.xlsx"
    write_xlsx([GRAMMAR_ROW], out)
    assert out.exists()


def test_write_xlsx_header(tmp_path):
    out = tmp_path / "grammar.xlsx"
    write_xlsx([], out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, 11)]
    assert headers == [
        "id", "name", "hsk_level", "description",
        "sentence_1", "translation_1",
        "sentence_2", "translation_2",
        "sentence_3", "translation_3",
    ]


def test_write_xlsx_data_row(tmp_path):
    out = tmp_path / "grammar.xlsx"
    write_xlsx([GRAMMAR_ROW], out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.cell(2, 1).value == "一01"
    assert ws.cell(2, 2).value == "方位名词：上、下、里"
    assert ws.cell(2, 3).value == "L1"
    assert ws.cell(2, 4).value in ("", None)
    assert ws.cell(2, 5).value == "书在桌子上。"
    assert ws.cell(2, 6).value in ("", None)


def test_write_xlsx_creates_parent_dir(tmp_path):
    out = tmp_path / "subdir" / "grammar.xlsx"
    write_xlsx([], out)
    assert out.exists()


def test_write_xlsx_empty_string_written_as_none_or_empty(tmp_path):
    """Empty string fields should be None or empty string in xlsx — not the string 'None'."""
    out = tmp_path / "grammar.xlsx"
    write_xlsx([GRAMMAR_ROW], out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    val = ws.cell(2, 4).value  # description
    assert val in ("", None)
