import sys
import pytest
import openpyxl
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from scripts.build_xlsx import fetch_json, merge_entries, load_level_entries, to_rows, write_xlsx, fetch_hskhsk_definitions, fetch_anki_definitions


ENTRY_A = {
    "simplified": "爱",
    "pos": ["verb"],
    "forms": [{"traditional": "愛", "transcriptions": {"pinyin": "ài"}, "meanings": ["to love", "to like"]}]
}
ENTRY_B = {
    "simplified": "本",
    "pos": ["noun"],
    "forms": [{"traditional": "本", "transcriptions": {"pinyin": "běn"}, "meanings": ["root", "origin"]}]
}
ENTRY_A_OLD = {
    "simplified": "爱",
    "pos": ["verb", "noun"],
    "forms": [{"traditional": "愛", "transcriptions": {"pinyin": "ài"}, "meanings": ["old definition"]}]
}


def test_fetch_json_returns_list():
    url = "https://raw.githubusercontent.com/drkameleon/complete-hsk-vocabulary/refs/heads/main/wordlists/inclusive/old/5.json"
    result = fetch_json(url)
    assert isinstance(result, list)
    assert len(result) > 0
    assert "simplified" in result[0]


def test_merge_entries_new_only():
    result = merge_entries({"爱": (ENTRY_A, 1)}, {})
    assert "爱" in result
    assert result["爱"]["source"] == "new-HSK L1"
    assert result["爱"]["entry"] == ENTRY_A


def test_merge_entries_old_only():
    result = merge_entries({}, {"本": (ENTRY_B, 2)})
    assert "本" in result
    assert result["本"]["source"] == "old-HSK L2"
    assert result["本"]["entry"] == ENTRY_B


def test_merge_entries_prefers_new_data():
    result = merge_entries({"爱": (ENTRY_A, 1)}, {"爱": (ENTRY_A_OLD, 3)})
    assert result["爱"]["entry"] == ENTRY_A
    assert result["爱"]["source"] == "old-HSK L3, new-HSK L1"


def test_to_rows_one_row_per_meaning():
    merged = merge_entries({"爱": (ENTRY_A, 1)}, {})
    rows = to_rows(merged)
    assert len(rows) == 2  # ENTRY_A has 2 meanings


def test_to_rows_columns():
    merged = merge_entries({"爱": (ENTRY_A, 1)}, {})
    row = to_rows(merged)[0]
    assert row["simplified"] == "爱"
    assert row["pinyin"] == "ài"
    assert row["traditional"] == "愛"
    assert row["pos"] == "verb"  # ENTRY_A has pos=["verb"] — not an abbreviated tag
    assert row["classifier"] == ""
    assert row["definition"] == "to love"
    assert row["source"] == "new-HSK L1"


def test_to_rows_classifier_joined():
    entry = {
        "simplified": "书",
        "pos": ["noun"],
        "forms": [{"traditional": "書", "transcriptions": {"pinyin": "shū"},
                   "meanings": ["book"], "classifiers": ["本", "册"]}]
    }
    merged = merge_entries({"书": (entry, 1)}, {})
    row = to_rows(merged)[0]
    assert row["classifier"] == "本, 册"


def test_to_rows_multiple_forms():
    entry = {
        "simplified": "好",
        "pos": ["adjective"],
        "forms": [
            {"traditional": "好", "transcriptions": {"pinyin": "hǎo"}, "meanings": ["good"]},
            {"traditional": "好", "transcriptions": {"pinyin": "hào"}, "meanings": ["to be fond of"]},
        ]
    }
    merged = merge_entries({"好": (entry, 1)}, {})
    rows = to_rows(merged)
    assert len(rows) == 2
    assert rows[0]["pinyin"] == "hǎo"
    assert rows[1]["pinyin"] == "hào"


def test_to_rows_pos_joined():
    entry = {
        "simplified": "爱",
        "pos": ["verb", "noun"],
        "forms": [{"traditional": "愛", "transcriptions": {"pinyin": "ài"}, "meanings": ["love"]}]
    }
    merged = merge_entries({"爱": (entry, 1)}, {})
    row = to_rows(merged)[0]
    assert row["pos"] == "verb, noun"  # not abbreviated tags, passed through as-is


def test_to_rows_uses_hskhsk_definitions():
    merged = merge_entries({"爱": (ENTRY_A, 1)}, {})
    hskhsk_defs = {"爱": ["love; affection", "to love"]}
    rows = to_rows(merged, hskhsk_defs)
    assert len(rows) == 2
    assert rows[0]["definition"] == "love; affection"
    assert rows[1]["definition"] == "to love"


def test_to_rows_falls_back_without_hskhsk():
    merged = merge_entries({"爱": (ENTRY_A, 1)}, {})
    rows = to_rows(merged)
    assert rows[0]["definition"] == "to love"


def test_to_rows_hskhsk_overrides_new_hsk():
    """Words in both old and new HSK use official definitions when available."""
    merged = merge_entries({"爱": (ENTRY_A, 1)}, {"爱": (ENTRY_A_OLD, 3)})
    hskhsk_defs = {"爱": ["official definition"]}
    rows = to_rows(merged, hskhsk_defs)
    assert len(rows) == 1
    assert rows[0]["definition"] == "official definition"


def test_to_rows_anki_fallback():
    """Words not in hskhsk_defs use anki_defs if available."""
    merged = merge_entries({"爱": (ENTRY_A, 1)}, {})
    rows = to_rows(merged, hskhsk_defs={}, anki_defs={"爱": ["anki definition"]})
    assert len(rows) == 1
    assert rows[0]["definition"] == "anki definition"


def test_to_rows_hskhsk_takes_priority_over_anki():
    merged = merge_entries({"爱": (ENTRY_A, 1)}, {})
    rows = to_rows(merged, hskhsk_defs={"爱": ["official"]}, anki_defs={"爱": ["anki"]})
    assert rows[0]["definition"] == "official"


def test_fetch_anki_definitions_parses_correctly(tmp_path):
    """Build a minimal .apkg and verify parsing."""
    import zipfile, sqlite3, json as json_mod

    db_path = tmp_path / "collection.anki2"
    conn = sqlite3.connect(db_path)
    conn.execute("CREATE TABLE notes (flds TEXT)")
    conn.execute("CREATE TABLE col (models TEXT)")
    flds = "\x1f".join([
        "爱",                  # word
        "",                    # interferences
        "<table><tbody><tr><th>{{c1::ài}}</th></tr>"
        "<tr><td>㊀ {{c1::to love}}<br>㊁ {{c1::to like}}</td></tr></tbody></table>",
        "", "", "", ""
    ])
    conn.execute("INSERT INTO notes VALUES (?)", (flds,))
    conn.execute("INSERT INTO col VALUES (?)", (json_mod.dumps({}),))
    conn.commit()
    conn.close()

    apkg_path = tmp_path / "test.apkg"
    with zipfile.ZipFile(apkg_path, "w") as z:
        z.write(db_path, "collection.anki2")

    result = fetch_anki_definitions(apkg_path)
    assert "爱" in result
    assert result["爱"] == ["to love", "to like"]


def test_merge_entries_both_present():
    result = merge_entries({"爱": (ENTRY_A, 1)}, {"爱": (ENTRY_A_OLD, 4), "本": (ENTRY_B, 2)})
    assert len(result) == 2
    assert result["爱"]["source"] == "old-HSK L4, new-HSK L1"
    assert result["本"]["source"] == "old-HSK L2"


ROW = {"simplified": "爱", "pinyin": "ài", "traditional": "愛",
       "pos": "verb", "classifier": "", "definition": "to love", "source": "new-HSK3"}


def test_write_xlsx_creates_file(tmp_path):
    out = tmp_path / "out.xlsx"
    write_xlsx([ROW], out)
    assert out.exists()


def test_write_xlsx_header(tmp_path):
    out = tmp_path / "out.xlsx"
    write_xlsx([], out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert [ws.cell(1, c).value for c in range(1, 8)] == [
        "simplified", "pinyin", "traditional", "pos", "classifier", "definition", "source"
    ]


def test_write_xlsx_data_row(tmp_path):
    out = tmp_path / "out.xlsx"
    write_xlsx([ROW], out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.cell(2, 1).value == "爱"
    assert ws.cell(2, 5).value in ("", None)
    assert ws.cell(2, 6).value == "to love"


def test_write_xlsx_creates_parent_dir(tmp_path):
    out = tmp_path / "subdir" / "out.xlsx"
    write_xlsx([], out)
    assert out.exists()


def test_write_xlsx_merges_same_word_rows(tmp_path):
    rows = [
        {"simplified": "爱", "pinyin": "ài", "traditional": "愛", "pos": "verb",
         "classifier": "", "definition": "to love", "source": "new-HSK3"},
        {"simplified": "爱", "pinyin": "ài", "traditional": "愛", "pos": "verb",
         "classifier": "", "definition": "to like", "source": "new-HSK3"},
    ]
    out = tmp_path / "out.xlsx"
    write_xlsx(rows, out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    merged_ranges = [str(r) for r in ws.merged_cells.ranges]
    # simplified (col 1) should be merged across rows 2-3
    assert "A2:A3" in merged_ranges
    # definition (col 6) should NOT be merged
    assert not any("F2" in r for r in merged_ranges)


def test_write_xlsx_no_merge_single_definition(tmp_path):
    out = tmp_path / "out.xlsx"
    write_xlsx([ROW], out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert len(list(ws.merged_cells.ranges)) == 0
